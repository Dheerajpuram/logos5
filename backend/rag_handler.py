
import os
import pandas as pd
import io
from PyPDF2 import PdfReader
import docx
import openpyxl
from langchain_google_genai import GoogleGenerativeAIEmbeddings, ChatGoogleGenerativeAI
from langchain_community.vectorstores import FAISS
from langchain_text_splitters import RecursiveCharacterTextSplitter
from utils import generate_forecast_plot

def _extract_csv_from_text(text: str) -> str:
    """Uses an LLM to find and format time-series data from raw text."""
    llm = ChatGoogleGenerativeAI(model="gemini-2.5-pro", google_api_key=os.getenv("GEMINI_API_KEY"))
    prompt = f"""
    You are an expert data extraction and cleaning assistant. Your primary task is to find and format any time-series data within the provided text into a strict CSV format with two columns: 'ds' for dates and 'y' for a numeric value.

    **Key Rules for Extraction:**
    1.  **Date Column ('ds')**: Dates must be in 'YYYY-MM-DD' format. If the original date has a different format (e.g., 'January 2023', '2023 Q1', '1/1/2023'), convert it to 'YYYY-MM-DD'. If only year or month/year is available, use the first day of the period (e.g., 'January 2023' becomes '2023-01-01').
    2.  **Value Column ('y')**: This column must contain a single numeric value corresponding to the date. Extract the most relevant quantitative measure that changes over time (e.g., sales, revenue, stock price, user count).
    3.  **Strict CSV Format**: The output must be pure CSV, starting with the header `ds,y` and then each row separated by a newline. Do NOT include any markdown code block fences (```csv) or other conversational text.
    4.  **No Time Data**: Do not include time components in the 'ds' column.
    5.  **Handling Missing Data**: If a date has multiple values, choose the most representative one (e.g., closing price, total sales). If no clear numeric value can be associated with a date, skip that entry.
    6.  **Prioritize Clear Time-Series**: Only extract data that clearly represents a series of values over distinct points in time.
    7.  **Return Empty if No Data**: If you cannot find any suitable time-series data, return an empty string. Do not guess or invent data.

    **Example (if text contains "In 2022, revenue was 100M. In 2023, it was 120M."):**
    ds,y
    2022-01-01,100000000
    2023-01-01,120000000

    **Example (if text contains "Sales: Jan $500, Feb $550, Mar $600"):**
    ds,y
    YYYY-01-01,500
    YYYY-02-01,550
    YYYY-03-01,600
    (Replace YYYY with the most probable year from context, or a placeholder if unknown)

    Text:
    ---
    {text}
    ---

    CSV Data:
    """
    try:
        response = llm.invoke(prompt)
        csv_data = response.content.strip()
        # The prompt explicitly asks not to include ```csv, but add a safeguard just in case
        if csv_data.startswith("```csv"):
            csv_data = csv_data[len("```csv"):].strip()
        if csv_data.endswith("```"):
            csv_data = csv_data[:-len("```")]
            
        # Basic validation: check for header and at least one data row
        if not csv_data or "\n" not in csv_data or not csv_data.startswith("ds,y"):
            print(f"--- LLM did not return valid CSV data: {csv_data[:100]}...")
            return ""
            
        return csv_data
    except Exception as e:
        print(f"--- Error during LLM data extraction: {e} ---")
        return ""

def answer_from_rag(query: str, selected_files: list[str] = [], plotting_intent: bool = False) -> dict:
    """Answers a question from documents, with an option to generate a forecast plot."""
    data_dir = "/Users/dheeraj/Desktop/finalmp/data"
    if not selected_files:
        return {"error": "Please select at least one file to query."}

    raw_text = ''
    for file_name in selected_files:
        file_path = os.path.join(data_dir, file_name)
        try:
            if file_path.endswith('.pdf'):
                with open(file_path, 'rb') as f:
                    reader = PdfReader(f)
                    for page in reader.pages:
                        raw_text += page.extract_text() + '\n'
            elif file_path.endswith('.docx'):
                doc = docx.Document(file_path)
                for para in doc.paragraphs:
                    raw_text += para.text + '\n'
            elif file_path.endswith('.xlsx'):
                workbook = openpyxl.load_workbook(file_path)
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    for row in sheet.iter_rows():
                        for cell in row:
                            if cell.value:
                                raw_text += str(cell.value) + ' '
                        raw_text += '\n'
        except Exception as e:
            return {"error": f"Error reading file {file_name}: {e}"}

    if not raw_text:
        return {"error": "No text could be extracted from the selected files."}

    # --- Forecasting Logic ---
    if plotting_intent:
        print("--- Plotting intent detected. Attempting to generate forecast... ---")
        csv_data = _extract_csv_from_text(raw_text)
        if not csv_data:
            return {"error": "Could not find or parse time-series data from the document(s) for forecasting."}
        
        df = pd.read_csv(io.StringIO(csv_data))
        forecast_result = generate_forecast_plot(df, selected_files[0])
        
        if "error" in forecast_result:
            return forecast_result # Pass the error up

        # Now, get a business analysis of the forecast
        llm = ChatGoogleGenerativeAI(model="gemini-2.5-pro", google_api_key=os.getenv("GEMINI_API_KEY"))
        analysis_prompt = f"""
        You are an expert business analyst. The following forecast has been generated from a document.
        Analyze the forecast summary and the original user query to provide a concise business analysis and interpretation.

        Original Query: "{query}"
        Forecast Summary: "{forecast_result['answer']}"

        Your concise analysis:
        """
        analysis_response = llm.invoke(analysis_prompt)
        forecast_result['answer'] = analysis_response.content.strip()
        return forecast_result

    # --- Standard RAG Logic ---
    else:
        text_splitter = RecursiveCharacterTextSplitter(chunk_size=1000, chunk_overlap=200)
        texts = text_splitter.split_text(raw_text)
        
        try:
            embeddings = GoogleGenerativeAIEmbeddings(model="models/embedding-001", google_api_key=os.getenv("GEMINI_API_KEY"))
            vectorstore = FAISS.from_texts(texts, embeddings)
            retriever = vectorstore.as_retriever()
            docs = retriever.get_relevant_documents(query)
        except Exception as e:
            return {"error": f"Error creating vector store or retrieving documents: {e}"}

        llm = ChatGoogleGenerativeAI(model="gemini-2.5-pro", google_api_key=os.getenv("GEMINI_API_KEY"))
        context = ' '.join([doc.page_content for doc in docs])
        prompt = f"""
        You are an expert business analyst. Your task is to analyze the provided context from documents and answer the user's query. When appropriate, provide insights, suggestions, and forecasts based on the data. If the documents do not contain enough information to make a forecast or suggestion, explain what information is missing.

        Context:
        {context}

        Query: "{query}"

        Based on the context, provide a concise and insightful answer to the query.
        """
        response = llm.invoke(prompt)
        return {"answer": response.content.strip()}

